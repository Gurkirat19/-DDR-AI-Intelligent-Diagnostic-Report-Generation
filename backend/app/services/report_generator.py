from pathlib import Path
from typing import Dict, Any
import json
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
import openpyxl


class ReportGenerator:
    def to_pdf(self, data: Dict[str, Any], output_path: Path) -> Path:
        c = canvas.Canvas(str(output_path), pagesize=LETTER)
        text = c.beginText(40, 750)
        text.textLine("AI DDR Report")
        for key, value in data.items():
            text.textLine(f"{key}: {value}")
        c.drawText(text)
        c.showPage()
        c.save()
        return output_path

    def to_json(self, data: Dict[str, Any], output_path: Path) -> Path:
        output_path.write_text(json.dumps(data, indent=2))
        return output_path

    def to_excel(self, data: Dict[str, Any], output_path: Path) -> Path:
        """Export DDR sections to Excel with headers for structured rows."""
        wb = openpyxl.Workbook()
        for sheet_name, rows in data.items():
            ws = wb.create_sheet(title=str(sheet_name)[:31])
            if isinstance(rows, list):
                # If list of dicts, add header row then aligned values
                if rows and all(isinstance(r, dict) for r in rows):
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    for row in rows:
                        ws.append([row.get(h) for h in headers])
                else:
                    for row in rows:
                        ws.append([row])
            elif isinstance(rows, dict):
                ws.append(list(rows.keys()))
                ws.append(list(rows.values()))
            else:
                ws.append([rows])

        # remove default sheet if unused
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        wb.save(output_path)
        return output_path
